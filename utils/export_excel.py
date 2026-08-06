"""
utils/export_excel.py
Builds a styled multi-sheet .xlsx workbook from session state.
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
import streamlit as st

# ── Colour palette matching SURM Excel ───────────────────────────────
GREEN_DARK   = "1F6B3A"
GREEN_LIGHT  = "C6EFCE"
YELLOW_FILL  = "FFFBE6"
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Calibri", size=10)
HEADER_FILL  = PatternFill("solid", fgColor=GREEN_DARK)
ALT_FILL     = PatternFill("solid", fgColor="F2F2F2")
THIN_BORDER  = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _write_df_to_sheet(ws, df: pd.DataFrame, title: str, start_row: int = 1):
    """Writes a DataFrame to a worksheet with SURM styling."""
    # Title row
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=max(len(df.columns), 1))
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    title_cell.fill  = PatternFill("solid", fgColor=GREEN_DARK)
    title_cell.alignment = CENTER

    if df.empty:
        ws.cell(row=start_row+1, column=1, value="(No data)").font = BODY_FONT
        return

    # Header row
    hr = start_row + 1
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=hr, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = PatternFill("solid", fgColor="2E7D52")
        cell.alignment = CENTER
        cell.border    = THIN_BORDER

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 1):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=hr + row_idx, column=col_idx, value=str(value) if value is not None else "")
            cell.font      = BODY_FONT
            cell.fill      = fill
            cell.alignment = LEFT
            cell.border    = THIN_BORDER

    # Auto column width
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *[len(str(v)) for v in df[col_name].astype(str)],
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Freeze header
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)


def build_excel_export() -> bytes:
    """
    Assembles the full SURM workbook from session state.
    Returns bytes for st.download_button().
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    ss = st.session_state

    # ── Sheet: Front Page ─────────────────────────────────────────────
    ws = wb.create_sheet("Front Page")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "SURM — Subsurface Uncertainty & Risk Management Plan"
    c.font  = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor=GREEN_DARK)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 36

    fields = [
        ("Project Name",  ss.get("project_name",  "")),
        ("Field Name",    ss.get("field_name",    "")),
        ("Project Phase", ss.get("project_phase", "")),
    ]
    for i, (label, val) in enumerate(fields, 3):
        ws.cell(row=i, column=1, value=label).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=i, column=2, value=val).font   = Font(name="Calibri", size=11)

    # Sign-off block
    signoff_headers = ["Role", "Name", "Date"]
    signoff_data    = [
        ["Prepared By",      ss.get("prep_name", ""),      ss.get("prep_date", "")],
        ["Reviewed By (G&G)", ss.get("rev_gg_name", ""),    ss.get("rev_gg_date", "")],
        ["Reviewed By (RE)",  ss.get("rev_re_name", ""),    ss.get("rev_re_date", "")],
        ["Reviewed By (PP)",  ss.get("rev_pp_name", ""),    ss.get("rev_pp_date", "")],
        ["Endorsed By",       ss.get("endorsed_name", ""), ss.get("endorsed_date", "")],
    ]
    ws.cell(row=7, column=1, value="Sign-Off").font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    ws.cell(row=7, column=1).fill = PatternFill("solid", fgColor=GREEN_DARK)
    ws.merge_cells("A7:C7")
    for ci, h in enumerate(signoff_headers, 1):
        cell = ws.cell(row=8, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="2E7D52")
    for ri, row in enumerate(signoff_data, 9):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val).border = THIN_BORDER

    # ── Sheet: Team ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Documentation")
    team_df = pd.DataFrame(ss.get("team_members", []))
    _write_df_to_sheet(ws2, team_df, "Team Members & Documentation")

    # ── Sheet: Tab 1 — Uncertainties ─────────────────────────────────
    ws3 = wb.create_sheet("1. Uncertainties List")
    unc_rows = [
        {"Discipline": u["discipline"], "Uncertainty": u["name"], "Selected": "Y" if u["selected"] else ""}
        for u in ss.get("uncertainties", [])
    ]
    _write_df_to_sheet(ws3, pd.DataFrame(unc_rows), "Uncertainties List")

    # ── Sheet: Tab 2 — Key Decisions ─────────────────────────────────
    ws4 = wb.create_sheet("2. Key Decisions")
    kd_df = pd.DataFrame(ss.get("key_decisions", []))
    _write_df_to_sheet(ws4, kd_df, "Key Project Decisions")

    # ── Sheet: Tab 3 — Impact Assessment ─────────────────────────────
    ws5 = wb.create_sheet("3. Impact Assessment")
    ia_df = pd.DataFrame(ss.get("impact_assessment", []))
    _write_df_to_sheet(ws5, ia_df, "Impact Assessment")

    # ── Sheet: Tab 4 — Key Uncertainties ─────────────────────────────
    ws6 = wb.create_sheet("4. Key Uncertainties")
    ku_df = pd.DataFrame(ss.get("key_uncertainties", []))
    _write_df_to_sheet(ws6, ku_df, "Key Uncertainties (Ranked)")

    # ── Sheet: Tab 5 — Resolution List ───────────────────────────────
    ws7 = wb.create_sheet("5. Resolution List")
    rl_data = ss.get("resolution_list", {})
    rl_rows = []
    for name, opts in rl_data.items():
        row = {"Uncertainty": name}
        row.update(opts)
        rl_rows.append(row)
    _write_df_to_sheet(ws7, pd.DataFrame(rl_rows) if rl_rows else pd.DataFrame(), "Resolution Alternatives")

    # ── Sheet: Tab 6 — Resolution Planner ────────────────────────────
    ws8 = wb.create_sheet("6. Resolution Planner")
    rp_df = pd.DataFrame(ss.get("resolution_planner", []))
    _write_df_to_sheet(ws8, rp_df, "Resolution Planner")

    # ── Sheet: Tab 7 — Risk Register ─────────────────────────────────
    ws9 = wb.create_sheet("7. Risk Register")
    rr_df = pd.DataFrame(ss.get("risk_register", []))
    _write_df_to_sheet(ws9, rr_df, "Risk Register")

    # ── Sheet: PRA Output ─────────────────────────────────────────────
    ws10 = wb.create_sheet("PRA Output")
    pra_df = pd.DataFrame(ss.get("pra_output", []))
    _write_df_to_sheet(ws10, pra_df, "PRA Output — Risk Register")

    # ── Save to bytes ─────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
